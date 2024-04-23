from openpyxl import load_workbook
from datetime import timedelta
from datetime import datetime
import datetime
import os
import time
import logging
from logging import handlers

"""
Below are some useful time operating functions
Mainly convert time date into string with different formats
"""
class TimeOPs(object):
    def __init__(self):
        self.time = datetime.datetime.now()

    def return_format_HH_colon_MM(self):
        return time.strftime("%H:%M")

    def return_format_YYYY_slash_MM_slash_DD(self):
        return time.strftime("%d/%m/%Y")

    def return_format_YYYYMMDD_hyphen_HHMM(self):
        return time.strftime("%Y%m%d-%H%M")

    def return_format_detail(self):
        return time.strftime("%c Week %W")
    def return_format_m(self):
        return time.strftime("m")
"""
Below are some useful file operating functions
"""
class FileOPs(object):
    def create_folder(sel, path):
        folder = os.path.exists(path)
        if not folder:
            os.mkdir(path)
        else:
            print("\n*Folder " + path + " has already existed please wait one minuet*\n")

    def delete_all_files_under(self, path):
        for i in os.listdir(path):
            file = path + "\\" + i
            if os.path.isfile(file):
                os.remove(file)
"""
Below are some useful log operating functions
"""
class LogOPs(object):
    def __init__(self, log_root, time, encode, console_level, file_level):
        self.logger = logging.getLogger(time.return_format_YYYYMMDD_hyphen_HHMM())
        self.logger.setLevel(file_level)
        file_path = log_root + "\\" + time.return_format_YYYYMMDD_hyphen_HHMM()
        ch = logging.StreamHandler()
        ch.setLevel(console_level)
        fh = logging.handlers.TimedRotatingFileHandler(filename=file_path, when='D', encoding=encode, backupCount=100)
        fh.setLevel(file_level)
        self.logger.addHandler(ch)
        self.logger.addHandler(fh)

        self.logger.info("Log Name: "+time.return_format_YYYYMMDD_hyphen_HHMM()+
                         "\nLog Time: "+time.return_format_detail()+
                         "\nLog Level: "+str(console_level)+"(console) "+str(file_level)+"(file)"+
                         "\nBack count: 100"+"\n")

    def get_logger(self):
        return self.logger

"""
Below are some useful Excel operating functions
"""
class ExcelOPs(object):
    def find_first_empty_cell_by_row(self, worksheet, based_row, col_low, col_high):
        for i in range(col_low, col_high):
            tmp = worksheet.cell(row=based_row, column=i)
            if tmp.value == None:
                return i
    def clear_cells_in_range(self,worksheet, row_low, row_high, col_low, col_high ):
        for i in range(row_low, row_high):
            for j in range(col_low, col_high):
                worksheet.cell(row=i, column=j).value = ''

    def load_original_workbook(self, excel_file_path):
        return load_workbook(excel_file_path)
    def get_worksheet(self, original_workbook, worksheet_name):
        activated_wb = original_workbook.active
        return original_workbook[worksheet_name]
    def save_workbook(self, original_workbook, save_as):
        original_workbook.save(save_as)
    def fill_bpu_speed_measure_result(self, worksheet1, result, curr_time):
        # find the first empty cell
        column_index = self.find_first_empty_cell_by_row(worksheet1, 9, 7, 70)

        # fill in the date, time and the measure result
        worksheet1.cell(row=9, column=column_index).value = curr_time.return_format_YYYY_slash_MM_slash_DD()
        worksheet1.cell(row=10, column=column_index).value = curr_time.return_format_HH_colon_MM()
        for i in range(11, 21):
            worksheet1.cell(row=i, column=column_index).value = result[i-11]

    def fill_monthly_bpu_speed_measure_record(self, worksheet1, worksheet2, result, curr_time):
        # count the times of speed measure within current month
        record_number = 0
        col_index = self.find_first_empty_cell_by_row(worksheet1, 9, 7, 70) - 1
        while(worksheet1.cell(row=9, column=col_index-record_number).value[3:5] == curr_time.return_format_m()):
            record_number += 1

        # figure out whether current week is the end of current month
        next_monday = curr_time.time + timedelta(days=7-curr_time.time.weekday())
        # print(record_number)
        if(next_monday.month == curr_time.time.month+1):
            # clear the last data of last month
            self.clear_cells_in_range(worksheet2,6,16,2,7)

            # fill in the data of current month
            for i in range(2, 2+record_number):
                for j in range(6, 16):
                    worksheet2.cell(row=j, column=i).value = worksheet1.cell(row=j+5, column=col_index-record_number-2+i).value
            for j in range(6, 16):
                worksheet2.cell(row=j, column=2+record_number).value = result[j - 6]


